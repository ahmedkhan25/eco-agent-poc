import os
import sys
import json
import urllib.parse
import requests
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# ---- CONFIGURE THESE ----
EXCEL_PATH = "olympia-weblinks.xlsx"  # Try this first
EXCEL_PATH_ALT = "Olympia Weblinks .xlsx"  # Alternative name
CSV_PATH = "Olympia Weblinks .xlsx - Olympia.csv"  # Fallback to CSV
OUTPUT_DIR = "pdfs"
INVENTORY_FILE = "inventory.json"  # Local inventory file

# If a cell is just "SLR-Plan-Final.pdf", we build:
#   BASE_PDF_URL + "SLR-Plan-Final.pdf"
# Adjust this to the real base path for Olympia PDFs.
BASE_PDF_URL = "https://www.olympiawa.gov/Document_center/"


def is_pdf_url(url: str) -> bool:
    """Check if URL points to a PDF - strict filtering to exclude HTML pages, databases, etc."""
    if not url:
        return False
    
    url_lower = url.lower()
    
    # Skip common non-PDF patterns
    skip_patterns = [
        'dashboard',
        'calendar',
        'database',
        '.html',
        '.htm',
        '/page/',
        '/view/',
        'city of olympia',  # Generic city pages
        'usda plants database',
        'municipal code',  # Usually HTML
    ]
    
    for pattern in skip_patterns:
        if pattern in url_lower:
            return False
    
    # Must end with .pdf or have .pdf in the path (before query params)
    parsed = urllib.parse.urlparse(url)
    path = parsed.path.lower()
    
    # Check if path ends with .pdf
    if path.endswith('.pdf'):
        return True
    
    # Check if .pdf appears in the path (but not as part of another word)
    if '.pdf' in path:
        # Make sure it's actually .pdf and not part of a longer extension
        parts = path.split('.pdf')
        if len(parts) > 1:
            # Check what comes after .pdf - should be nothing or query params
            after_pdf = parts[1]
            if not after_pdf or after_pdf.startswith('/') or after_pdf.startswith('?'):
                return True
    
    return False


def find_excel_file():
    """Find the Excel file in the current directory"""
    if os.path.exists(EXCEL_PATH):
        return EXCEL_PATH
    if os.path.exists(EXCEL_PATH_ALT):
        return EXCEL_PATH_ALT
    return None


def extract_hyperlinks_from_excel(file_path):
    """Extract hyperlinks from Excel file using openpyxl"""
    try:
        wb = load_workbook(file_path, data_only=False)
        ws = wb.active
        
        links = []
        # Iterate through rows (assuming first column is title, second is link)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if len(row) < 2:
                continue
            
            title_cell = row[0]
            link_cell = row[1]
            
            title = str(title_cell.value).strip() if title_cell.value else ""
            
            # Try to get hyperlink from the cell
            url = None
            if link_cell.hyperlink:
                url = link_cell.hyperlink.target
            elif link_cell.value:
                cell_value = str(link_cell.value).strip()
                # Check if it's already a URL
                if cell_value.lower().startswith("http"):
                    url = cell_value
                elif cell_value.lower().endswith(".pdf"):
                    # Try to build URL from filename
                    url = urllib.parse.urljoin(BASE_PDF_URL, cell_value)
            
            # Only add if it's a PDF URL
            if url and is_pdf_url(url):
                links.append((title, url))
        
        return links
    except Exception as e:
        print(f"Error reading Excel file with openpyxl: {e}")
        return None


def read_from_pandas(file_path):
    """Read Excel file using pandas (fallback method)"""
    try:
        df = pd.read_excel(file_path)
        links = []
        
        # Assume first column is title, second is link
        for _, row in df.iterrows():
            if len(row) < 2:
                continue
            
            title = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            cell_value = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            
            if not cell_value:
                continue
            
            # Check if it's already a URL
            url = None
            if cell_value.lower().startswith("http"):
                url = cell_value
            elif cell_value.lower().endswith(".pdf"):
                url = urllib.parse.urljoin(BASE_PDF_URL, cell_value)
            
            # Only add if it's a PDF URL
            if url and is_pdf_url(url):
                links.append((title, url))
        
        return links
    except Exception as e:
        print(f"Error reading Excel file with pandas: {e}")
        return None


def read_from_csv(file_path):
    """Fallback: read from CSV file"""
    import csv
    links = []
    
    try:
        with open(file_path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            for row in reader:
                if len(row) < 2:
                    continue
                
                title = row[0].strip()
                cell = row[1].strip()
                
                if not cell:
                    continue
                
                url = None
                if cell.lower().startswith("http"):
                    url = cell
                elif cell.lower().endswith(".pdf"):
                    url = urllib.parse.urljoin(BASE_PDF_URL, cell)
                
                # Only add if it's a PDF URL
                if url and is_pdf_url(url):
                    links.append((title, url))
    except Exception as e:
        print(f"Error reading CSV file: {e}")
    
    return links


def load_inventory():
    """Load existing inventory or create new one"""
    if os.path.exists(INVENTORY_FILE):
        try:
            with open(INVENTORY_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Warning: Could not load inventory file: {e}")
            print("Creating new inventory...")
    
    # Create new inventory structure
    return {
        "metadata": {
            "created_at": datetime.utcnow().isoformat() + "Z",
            "updated_at": datetime.utcnow().isoformat() + "Z",
            "source_file": None,
            "bucket": "olympia-plans-raw"
        },
        "files": []
    }


def save_inventory(inventory):
    """Save inventory to JSON file"""
    inventory["metadata"]["updated_at"] = datetime.utcnow().isoformat() + "Z"
    with open(INVENTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(inventory, f, indent=2, ensure_ascii=False)


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # Load or create inventory
    inventory = load_inventory()
    
    # Try to find and read Excel file
    excel_file = find_excel_file()
    source_file = excel_file if excel_file else CSV_PATH
    links = []
    
    if excel_file:
        print(f"Reading Excel file: {excel_file}")
        # Try openpyxl first (better for hyperlinks)
        links = extract_hyperlinks_from_excel(excel_file)
        
        # Fallback to pandas if openpyxl didn't work
        if not links:
            print("Trying pandas method...")
            links = read_from_pandas(excel_file)
    else:
        # Fallback to CSV
        if os.path.exists(CSV_PATH):
            print(f"Reading CSV file: {CSV_PATH}")
            links = read_from_csv(CSV_PATH)
        else:
            print(f"Error: Could not find Excel file ({EXCEL_PATH} or {EXCEL_PATH_ALT}) or CSV file ({CSV_PATH})")
            sys.exit(1)
    
    if not links:
        print("No links found in file!")
        sys.exit(1)
    
    # Update inventory metadata
    inventory["metadata"]["source_file"] = source_file
    
    print(f"Found {len(links)} potential PDF links\n")
    
    downloaded = 0
    skipped = []
    
    # Create a map of existing files in inventory by filename
    existing_files = {f["local_filename"]: f for f in inventory["files"]}
    
    for title, url in links:
        if not is_pdf_url(url):
            skipped.append((title, url))
            continue
        
        filename = os.path.basename(urllib.parse.urlparse(url).path)
        # Decode URL-encoded filename
        filename = urllib.parse.unquote(filename)
        out_path = os.path.join(OUTPUT_DIR, filename)
        
        # Check if file already exists locally
        if os.path.exists(out_path):
            print(f"[skip] already exists: {filename}")
            # Update inventory if not already tracked
            if filename not in existing_files:
                file_info = {
                    "excel_title": title,
                    "source_url": url,
                    "local_filename": filename,
                    "s3_key": None,
                    "s3_bucket": inventory["metadata"]["bucket"],
                    "file_size": os.path.getsize(out_path),
                    "downloaded_at": datetime.utcnow().isoformat() + "Z",
                    "uploaded_at": None
                }
                inventory["files"].append(file_info)
                existing_files[filename] = file_info
            downloaded += 1
            continue
        
        print(f"[download] {title} -> {filename}")
        print(f"  URL: {url}")
        try:
            resp = requests.get(url, timeout=60, allow_redirects=True)
            resp.raise_for_status()
            
            # Verify it's actually a PDF by checking Content-Type
            content_type = resp.headers.get('content-type', '').lower()
            if 'pdf' not in content_type:
                print(f"  !! Skipping: Content-Type is {content_type}, not PDF")
                skipped.append((title, url))
                continue
            
            # Also check first few bytes for PDF magic number (%PDF)
            if len(resp.content) > 4:
                if not resp.content[:4].startswith(b'%PDF'):
                    print(f"  !! Skipping: File doesn't start with PDF magic number")
                    skipped.append((title, url))
                    continue
            
        except Exception as e:
            print(f"  !! error fetching {url}: {e}")
            skipped.append((title, url))
            continue
        
        # Save file
        with open(out_path, "wb") as out_f:
            out_f.write(resp.content)
        
        file_size = len(resp.content)
        downloaded_at = datetime.utcnow().isoformat() + "Z"
        
        # Add to inventory
        file_info = {
            "excel_title": title,
            "source_url": url,
            "local_filename": filename,
            "s3_key": None,  # Will be set during upload
            "s3_bucket": inventory["metadata"]["bucket"],
            "file_size": file_size,
            "downloaded_at": downloaded_at,
            "uploaded_at": None
        }
        
        # Update if exists, otherwise append
        if filename in existing_files:
            idx = inventory["files"].index(existing_files[filename])
            inventory["files"][idx].update(file_info)
        else:
            inventory["files"].append(file_info)
            existing_files[filename] = file_info
        
        downloaded += 1
        print(f"  ✓ Downloaded ({file_size} bytes)\n")
        
        # Save inventory after each download
        save_inventory(inventory)
    
    # Final save of inventory
    save_inventory(inventory)
    
    print(f"\n{'='*60}")
    print(f"Done. Downloaded {downloaded} PDFs.")
    print(f"Inventory saved to: {INVENTORY_FILE}")
    print(f"Total files in inventory: {len(inventory['files'])}")
    if skipped:
        print(f"\nSkipped {len(skipped)} rows (not PDFs or failed URLs):")
        for title, url in skipped[:10]:  # Show first 10
            print(f"  - {title!r} -> {url!r}")
        if len(skipped) > 10:
            print(f"  ... and {len(skipped) - 10} more")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
